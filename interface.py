# coding=utf-8
import os
import sys
import time
import xlrd
import datetime
from openpyxl import Workbook

reload(sys)
sys.setdefaultencoding('utf8')

from db import DBConn
# from werkzeug.security import generate_password_hash, check_password_hash

'''
table1_map = {
    '任单编号': 'any_single_num',
    '样品编号': 'sample_number',
    '样品名称': 'sample_name',
    '文库名称': 'library_name',
    'Index序号': 'index_num',
    'Index序列': 'index_sequence',
    '文库类型': 'library_type',
    '文库切胶长度': 'length_of_gel',
    '片段长度(bp)': 'fragment_length',
    '文库体积(ul)': 'library_volume',
    '数据量（raw data）': 'data_size'
}
'''


def save_info(all_info, username, action='new'):
    db_instance = DBConn()
    try:
        project_id = save_sample_project_master_info(db_instance, all_info.get('sample_project_master_info'), username, action)
        if action == 'new':
            save_log_info(db_instance, username, project_id, msg='create new project')

    except Exception, e:
        print e
        return {'data': '', 'errcode': 1, 'msg': '保存失败！项目编号已经存在，请另外选择项目编号。'}

    del all_info['sample_project_master_info']
    for key in all_info.keys():
        table_name = key.replace('_info', '')
        data_info = all_info.get(key)
        data_info['project_id'] = int(project_id)
        if action == 'update':
             db_instance.update(table_name, {'project_id': project_id}, data_info)
        else:
             db_instance.insert(table_name, data_info)

    msg = '更新成功！' if action == 'update' else '保存成功!'
    return {'data': project_id, 'errcode': 0, 'msg': msg}


def get_project_id_by_num(db_instance, project_number):
    cmd = "select id from sample_project_master where project_number=%s" % project_number
    result = db_instance.execute(cmd, get_all=False)

    return result[0]

'''
def get_project_log_by_num(db_instance, project_id):
    cmd = "select action from project_log_table where id=%s" % project_id
    result = db_instance.execute(cmd, get_all=False)

    return result[0] or ''
'''

def save_compare_text(project_id, all_info, current_time):
    try:
        db = DBConn()
        result = db.execute("select project_name from sample_project_master where id=%s" % project_id, get_all=False)
        project_name = result[0]
        project_dir = os.path.join(os.path.dirname(__file__), 'static/export/' + project_name)
        if not os.path.exists(project_dir):
            os.mkdir(project_dir)
        group_file_name = os.path.join(project_dir, 'sample_inf' + current_time + '.txt')
        compare_file_name = os.path.join(project_dir, 'group_compare' + current_time + '.txt')
        sample_packet_information = all_info['sample_packet_information']
        compare_table = all_info['compare_table']
        fd_group_file = open(group_file_name, 'w+')
        fd_compare_file = open(compare_file_name, 'w+')
        fd_group_file.write(all_info['reference_genome'] + '\n')
        for row in sample_packet_information:
            fd_group_file.write(row['sample_group'] + '\t' + row['sample_name'] + '\n')

        for row in compare_table:
            fd_compare_file.write(row['comparison_name'] + '\t' + row['sample_group1'] + '\t' + row['sample_group2'] + '\n')
        fd_group_file.close()
        fd_compare_file.close()
    except Exception, e:
        print e


def save_compare_input(all_info, username, project_id, action='new'):
    current_time = datetime.datetime.now()
    save_compare_text(project_id, all_info, str(current_time))
    db_instance = DBConn()
    #project_log = get_project_log_by_num(db_instance, project_id)
    all_info['project_id'] = project_id
    cmd = "select id from analysis_master where created_by='%s' and project_id=%s" % (username, project_id)
    result = db_instance.execute(cmd, get_all=False)
    if result:
        action = 'update'
        master_id = result[0]
        db_instance.delete('sample_packet_information', {'master_id': master_id})
        db_instance.delete('compare_table', {'master_id': master_id})
    sample_packet_information = all_info['sample_packet_information']
    compare_table = all_info['compare_table']
    del all_info['sample_packet_information']
    del all_info['compare_table']
    all_info['updated_by'] = username
    all_info['update_time'] = current_time
    if action == 'new':
        all_info['create_time'] = current_time
        all_info['created_by'] = username
        master_id = db_instance.insert('analysis_master', all_info)
        save_log_info(db_instance, username, project_id, msg='created compare method')
        #project_log += '\n%s: %s created compare method.\n' % (current_time, username)
        #db_instance.update('sample_project_master', {'id': project_id}, {'project_log': project_log})
    else:
        save_log_info(db_instance, username, project_id, msg='update compare method')
        #project_log += '\n%s: %s update compare method.\n' % (current_time, username)
        #db_instance.update('sample_project_master', {'id': project_id}, {'project_log': project_log})
        db_instance.update('analysis_master', {'id': master_id}, all_info)
    for row in sample_packet_information:
        del row['id']
        row['master_id'] = master_id
        db_instance.insert('sample_packet_information', row)

    for row in compare_table:
        del row['id']
        row['master_id'] = master_id
        db_instance.insert('compare_table', row)

    msg = '更新成功！' if action == 'update' else '保存成功!'
    return {'data': '', 'errcode': 0, 'msg': msg}


def save_table_info(db_instance, table_data, project_id, action):
    if action == 'update':
        cmd = "delete from sample_info_detail where project_id=%s" % project_id
        db_instance.execute(cmd)
    for row in table_data:
        del row['id']
        row['project_id'] = project_id
        if not row.get('sample_name'):
            continue
        else:
            db_instance.insert('sample_info_detail', row)


def save_sample_project_master_info(db_instance, data_info, username, action):
    if action == 'update':
        cmd = "select id from sample_project_master where project_number='%s'" % data_info['project_number']
        result = db_instance.execute(cmd, get_all=False)
        project_id = result[0]
        #data_info['project_log'] += "\n%s: %s update this project.\n" % (time, username)
        data_info['project_leader'] = username
        db_instance.update('sample_project_master', {'id': project_id}, data_info)
        save_log_info(db_instance, username, project_id, msg='update this project')
        return project_id
    else:
        cmd = "select max(project_number) from sample_project_master"
        result = db_instance.execute(cmd, get_all=False)
        name_list = result[0].split('-')
        new_project_num = name_list[0] + '-' + str(int(name_list[1]) + 1)
        data_info['project_number'] = new_project_num
        data_info['created_by'] = username
        data_info['create_time'] = datetime.datetime.now()
        #data_info['project_log'] = "%s: %s create new project.\n" % (time, username)
        return db_instance.insert('sample_project_master', data_info)


def save_log_info(db_instance, username, project_id, msg):
    time = datetime.datetime.now()
    insert_dict = dict(project_id=project_id, action=msg,
                           manager=username, time=time)
    db_instance.insert('project_log_table', insert_dict)

def show_all_data(username, role='user'):
    data = []
    cmd = """select * from sample_project_master spm,
            sample_species ss,
            sample_type st,
            dna_sample_sequencing_type dsst,
            rna_sample_sequencing_type rsst
            where ss.project_id=spm.id and
            st.project_id=spm.id and
            dsst.project_id=spm.id and
            rsst.project_id=spm.id"""
    if role == 'manager':
        cmd += " and spm.project_leader='%s'" % username
    elif role == 'user':
        cmd += " and spm.created_by='%s'" % username

    db = DBConn()
    result = db.execute(cmd)
    for i in result:
        temp_dict = dict(i)
        data.append(temp_dict)

    return data


def get_all_user_data():
    data = []
    cmd = "select username,e_mail email,tel,company,age,role,field,status,notes from user_info"
    results = DBConn().execute(cmd)
    for result in results:
        data.append(dict(result))

    return data


def get_one_project_data(project_id):
    cmd = """select * from sample_project_master spm,
             sample_species ss,
             sample_type st,
             dna_sample_sequencing_type dsst,
             rna_sample_sequencing_type rsst,
             sample_other other
             where ss.project_id=spm.id and
             st.project_id=spm.id and
             dsst.project_id=spm.id and
             rsst.project_id=spm.id and
             other.project_id=spm.id and
             spm.id=%s""" % project_id
    db = DBConn()
    result = db.execute(cmd, get_all=False)
    data = dict(result)
    return data


def get_analysis_data(username, role, project_id):

    if not project_id:
        return {}
    cmd = """select * from analysis_master where project_id=%s""" % project_id
    db = DBConn()
    result = db.execute(cmd, get_all=False)
    return dict(result) if result else {}


def phone_check(s):
    phone_prefix = ['130', '131', '132', '133', '134', '135', '136', '137', '138', '139', '150',
                    ' 151', '152', '153', '156', '158', '159', '170', '183', '182', '185', '186', '188', '189']

    return True if len(s) == 11 and s.isdigit() and s[:3] in phone_prefix else False


# def get_user_role(username, password=''):
#     db = DBConn()
#     if phone_check(username):
#         cmd = "select role,status from user_info where tel='%s'" % username
#     else:
#         cmd = "select role,status from user_info where username='%s'" % username
#     if password:
#         cmd += " and password='%s'" % password
#     result = db.execute(cmd, get_all=False)
#     role = result[0] if result else ''
#     status = result[1] if result else ''
#     return role, status


def get_user_role(username, password=''):
    db = DBConn()
    if phone_check(username):
        cmd = "select role,status,password from user_info where tel='%s'" % username
    else:
        cmd = "select role,status,password from user_info where username='%s'" % username
    result = db.execute(cmd, get_all=False)
    password_hash = result[2] if result else ''
    if password:
        if password_hash == password:
            role = result[0] if result else ''
            status = result[1] if result else ''
            return role, status
        else:
            return '', ''
    else:
        role = result[0] if result else ''
        status = result[1] if result else ''
        return role, status


def get_other_info(username):
    db = DBConn()
    if phone_check(username):
        cmd = "select e_mail, tel, company, field, customer_name from user_info where tel='%s'" % username
    else:
        cmd = "select e_mail, tel, company, field, customer_name from user_info where username='%s'" % username
    result = db.execute(cmd, get_all=False)

    return result if result else ('', '', '', '')


def modify_base_info(info):
    time = datetime.datetime.now()
    info['update_time'] = time
    username = info.get('username')
    tel = info.get('tel')
    db = DBConn()
    cmd = "select id from user_info where username!='%s' and tel='%s'" % (username, tel)
    check_result = db.execute(cmd, get_all=False)
    if check_result:
        return {'data': '', 'errcode': 1, 'msg': '该电话号码已经被其它用户注册，请使用其它号码！'}
    else:
        db.update('user_info', {'username': username}, info)
        return {'data': '', 'errcode': 0, 'msg': '更新成功！'}


def change_password(info):
    username = info['username']
    db = DBConn()
    cmd = "select password from user_info where username='%s'" % username
    result = db.execute(cmd, get_all=False)
    if result and result[0] == info['old_passwd'] and info['new_passwd'] != info['old_passwd']:
        # password_hash = generate_password_hash(info['new_passwd'])
        db.update('user_info', {'username': username}, {'password': info['new_passwd'],
                                                        'update_time': datetime.datetime.now()})

        return {'data': '', 'errcode': 0, 'msg': '更新成功！'}
    elif result and result[0] != info['old_passwd']:
        return {'data': '', 'errcode': 1, 'msg': '旧密码错误！'}
    elif info['new_passwd'] == info['old_passwd']:
        return {'data': '', 'errcode': 1, 'msg': '新密码和旧密码不能一样！'}
    else:
        return {'data': '', 'errcode': 1, 'msg': '用户名不存在！'}


def save_user_info(info):
    time = datetime.datetime.now()
    info['create_time'] = time
    info['update_time'] = time
    info['status'] = 'R'  # 'R' means review
    if len(info['password']) < 5:
        return {'data': '', 'errcode': 3, 'msg': "密码设置过短，请设置为5位以上！"}
    db = DBConn()
    name_result = db.execute("select id, status from user_info where username='%s'" % info.get('username'), get_all=False)
    tel_result = db.execute("select id, status from user_info where tel='%s'" % info.get('tel'), get_all=False)

    def fail_info(sql_result, check_type):
        status = sql_result[1]
        if status == 'R':
            msg = '该{check_type}之前已经注册，但状态处于等待中，请使用其它{check_type}注册！'
        elif status == 'Y':
            msg = '注册失败：该{check_type}已经注册，请使用其它{check_type}！'
        else:
            msg = '注册失败：该{check_type}已经存在！'

        return {'data': '', 'errcode': 2, 'msg': msg.format(check_type=check_type)}

    if name_result:
        ret = fail_info(name_result, '名字')
    elif tel_result:
        ret = fail_info(tel_result, '电话号码')
    else:
        # info['password'] = generate_password_hash(info.get('password'))
        db.insert('user_info', info)
        ret = {'data': '', 'errcode': 0, 'msg': '注册成功：等待审核！'}

    return ret


def get_detail_sample_data(project_number):
    data = []
    cmd = """SELECT d.* FROM sample_info_detail d, sample_project_master m
            where m.id=d.project_id and m.project_number=%s""" % project_number
    db = DBConn()
    results = db.execute(cmd)
    for result in results:
        data.append(dict(result))

    return data


def get_analysis_table_data(project_id):
    data = {}
    db = DBConn()
    cmd = """select info.* from sample_packet_information info, analysis_master m
                  where m.id=info.master_id and m.project_id=%s""" % project_id
    results = db.execute(cmd)
    data['sample_packet_information'] = [dict(i) for i in results]
    cmd = """select info.* from compare_table info, analysis_master m
                  where m.id=info.master_id and m.project_id=%s""" % project_id
    results = db.execute(cmd)
    data['compare_table'] = [dict(i) for i in results]

    return data


def save_status(project_number, status):
    db = DBConn()
    db.update('sample_project_master', {'project_number': project_number}, {'status': status})


def admin_save_user_info(username, status):
    db = DBConn()
    db.update('user_info', {'username': username}, {'status': status})


def get_project_number_list(username, user_role):
    if user_role == 'manager':
        cmd = "select id, concat(project_name, '-', project_number) from sample_project_master where project_leader='%s'" % username
    elif user_role == 'user':
        cmd = "select id, concat(project_name, '-', project_number) from sample_project_master where created_by='%s'" % username

    db = DBConn()
    results = db.execute(cmd)
    data = {}
    for result in results:
        data[result[0]] = result[1]

    return data


def get_manager_list():
    db = DBConn()
    cmd = "select username from user_info where role='manager'"

    result = db.execute(cmd)

    return [i[0] for i in result]


def transfer_table_title(table_title):
    title_map = {
        '任单编号': 'any_single_num',
        '样品编号': 'sample_number',
        '样品名称': 'sample_name',
        '文库名称': 'library_name',
        'Index序号': 'index_num',
        'Index序列': 'index_sequence',
        '文库类型': 'library_type',
        '文库切胶长度': 'length_of_gel',
        '片段长度(bp)': 'fragment_length',
        '文库体积(ul)': 'library_volume',
        '数据量（raw data）': 'data_size',

        'WGCID': 'wg_cid',
        'LibID': 'lib_id',
        'SampleType': 'sample_type',
        'qRCB': 'q_rcb',
        'volume': 'volume',
        'OD': 'od',
        'RIN': 'rin',
        'LibSize': 'lib_size',
        'Quality': 'qty',

        'Original Sample Name': 'original_sample_name',
        'Project ID': 'project_id_e',
        'Yield': 'yield',
        'Reads': 'reads'

    }
    for index, title in enumerate(table_title):
        if '数据量' in title:
            title = '数据量（raw data）'

        if str(title) in title_map.keys():
            table_title[index] = title_map[str(title)]

    return table_title


def transfer_excel_to_json(file_name):
    # file_name = os.path.basename(file_name)
    data = xlrd.open_workbook(file_name)
    sheets = data.sheet_names()
    json_data = {}
    for sheet in sheets:
        table = data.sheet_by_name(sheet)
        n_rows = table.nrows
        n_cols = table.ncols
        table_data = []
        table_title = table.row_values(0)
        table_title = transfer_table_title(table_title)
        for i in range(1, n_rows):
            table_data.append(dict(zip(table_title, table.row_values(i))))
        json_data[sheet] = table_data

    return json_data


def upload_excel(io_stream):
    """
    Upload file and write it to database
    :param filename: filename
    :param io_stream: IO Stream
    :return: dict information
    """
    upload_path = os.path.dirname(__file__) + '/static/import/'
    file_path = os.path.join(upload_path, 'temp.xls')
    #juge file whether has exist
    if os.path.exists(file_path):
        os.remove(file_path)
    temp_file_path = file_path + '~'
    output_file = open(temp_file_path, 'wb')
    io_stream.seek(0)
    while True:
        data = io_stream.read(2 << 16)
        if not data:
            break
        output_file.write(data)

    output_file.flush()
    output_file.close()
    os.rename(temp_file_path, file_path)
    data = transfer_excel_to_json(file_path)
    return {'data': data, 'errcode': 0, 'msg': '上传成功'}


def call_sms_service(ipaddr):
    import requests
    account_sid = '518d949550f3567866c2694f4988f889'
    auth_token = 'ca4624f01d3ed40b30d09ff1cc1c374b'
    service_host = 'https://api.ucpaas.com/2014-06-30/Accounts/{account_sid}/Messages/templateSMS'
    payload = {
         "templateSMS": {
             "appId": "caabd9b5be1b4ff8a8592068b230c339",
             "param": "0000",
             "templateId": "30213",
             "to": ipaddr
             }
    }

    head = {
        'Accept': 'application/json',
        'Content-Type': 'application/json;charset=utf-8',
        'Authorization': 'ZTAzYmM5MTA2YzZlZDBlYWViZmNlOGMzNjhmZGNkNDg6MjAxNDA2MjMxODUwMjE'
    }

    r = requests.post(service_host, data=payload, headers=head)
    print r
    return


def export_user_info():
    book = Workbook()
    sheet1 = book.worksheets[0]
    all_user_data = get_all_user_data()
    title_map = {'用户名': 'username',
                 '电子邮件': 'email',
                 '手机号码': 'tel',
                 '客户单位': 'company',
                 '年龄': 'age',
                 '角色': 'role',
                 '研究领域': 'field',
                 '状态': 'status',
                 '备注': 'notes'}
    title_list = ['用户名', '电子邮件', '手机号码', '客户单位', '年龄', '角色', '研究领域', '状态']
    for i, title in enumerate(title_list):
        sheet1.cell(row=1, column=i + 1).value = title
    for i, data in enumerate(all_user_data):
        for j, title in enumerate(title_list):
            sheet1.cell(row=i+2, column=j + 1).value = data.get(title_map[title])
    export_path = os.path.dirname(__file__) + '/static/export/'
    file_name = 'user_info_' + datetime.datetime.now().strftime("%Y%m%d%H%M%S") + '.xls'
    full_file_name = os.path.join(export_path, file_name)
    # remove old user info table
    for name in os.listdir(export_path):
        if 'user_info' in name:
            os.remove(os.path.join(export_path, name))
    book.save(filename=full_file_name)

    return file_name


def upload_project_file(io_stream, filename, project_number, project_name):
    filename = filename.split('\\')[-1]
    project_folder = project_number + '-' + project_name
    upload_path = os.path.join(os.path.dirname(__file__), 'static/import/', project_folder)
    if not os.path.exists(upload_path):
        os.makedirs(upload_path)
    file_path = os.path.join(upload_path, filename)
    # juge file whether has exist
    if os.path.exists(file_path):
        os.remove(file_path)
        print "delete had exist file: %s" % filename
    temp_file_path = file_path + '~'
    output_file = open(temp_file_path, 'wb')
    io_stream.seek(0)
    while True:
        data = io_stream.read(2 << 16)
        if not data:
            break
        output_file.write(data)
    output_file.flush()
    output_file.close()
    os.rename(temp_file_path, file_path)

    return {'data': '', 'errcode': 0, 'msg': "upload ok!"}


def get_project_files(project_number, project_name):
    file_list = []
    project_folder = project_number + '-' + project_name
    project_folder_path = os.path.join(os.path.dirname(__file__), 'static/import/', project_folder)
    if not os.path.exists(project_folder_path):
        return {'data': [], 'errcode': 0, 'msg': ""}
    files = os.listdir(project_folder_path)
    for file_name in files:
        file_path = os.path.join(project_folder_path, file_name)
        stat_info = os.stat(file_path)
        file_list.append({
            'file_name': file_name,
            'create_time': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat_info.st_ctime)),
            'update_time': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(stat_info.st_mtime))
        })

    return {'data': file_list, 'errcode': 0, 'msg': ""}


def save_sample_data(project_id, data, username):
    try:
        table_name = "sample_table"
        db = DBConn()
        result = db.execute("select * from %s where project_id='%s'" %(table_name,project_id))
        if result:
            action = 'update'
        else:
            action = 'new'
        db.delete(table_name, {'project_id': project_id})
        for row in data:
            row['project_id'] = project_id
            db.insert(table_name, row)
        if action == 'update':
            save_log_info(db, username, project_id, msg='update sample table')
        else:
            save_log_info(db, username, project_id, msg='create sample table')
        ret = {'data': '', 'errcode': 0, 'msg': ''}
    except Exception, e:
        import traceback
        traceback.print_exc()
        ret = {'data': '', 'errcode': 1, 'msg': '%s DB issue!' % e}

    return ret


def get_sample_list_by_project(project_id):
    sample_list = {}
    db = DBConn()
    cmd = "SELECT id_alias,sample_name FROM SEQ_SA_INFO.sample_info_detail where project_id=%s" % project_id
    results = db.execute(cmd)
    for result in results:
        sample_list[result[0]] = result[1]

    return sample_list


def get_sample_table_data(project_id):
    all_project_sample_data = []
    db = DBConn()
    cmd = "SELECT * FROM SEQ_SA_INFO.sample_table where project_id=%s" % project_id
    results = db.execute(cmd)
    for result in results:
        all_project_sample_data.append(dict(result))

    return all_project_sample_data


def get_project_info(project_id):
    db = DBConn()
    cmd = "SELECT project_name, project_number FROM sample_project_master where id=%s" % project_id
    result = db.execute(cmd, get_all=False)

    return dict(result)


'''
add by chencheng
'''


def get_project_leader():
    db = DBConn()
    cmd = "SELECT username FROM user_info where role='manager'"
    results = db.execute(cmd)
    project_leaders = []
    for result in results:
        project_leaders.append(result[0])

    return project_leaders


def get_log_data(project_id):
    data = []
    cmd = "SELECT * FROM project_log_table where project_id='%s'" % project_id

    db = DBConn()
    results = db.execute(cmd)
    for result in results:
        data.append(dict(result))

    return data


if __name__ == '__main__':
    print get_project_files('111', '111')
